"""Seed the initial_state sheet in the input Excel workbook.

Reads the existing workbook, computes a max-expansion warm-start, and appends
(or replaces) the ``initial_state`` sheet so the user can hand-edit it.

Usage:
    python scripts/seed_initial_state.py
"""
from __future__ import annotations

import os
import sys

# Allow running from repo root
sys.path.insert(0, os.path.join(os.path.dirname(__file__), "..", "model"))

import pandas as pd
from data_prep import load_data_from_excel

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
PROJECT_ROOT = os.path.dirname(SCRIPT_DIR)
EXCEL_PATH = os.path.join(PROJECT_ROOT, "inputs", "input_data_intertemporal.xlsx")

_TIMES = ["2025", "2030", "2035", "2040", "2045"]


def main() -> None:
    data = load_data_from_excel(EXCEL_PATH)
    times = data.times or list(_TIMES)
    regions = data.regions
    ytn_dict = data.years_to_next or {t: 5.0 for t in times}
    g_exp_dict = data.g_exp_ub or {r: 0.0 for r in regions}
    g_exp_is_abs = bool(getattr(data, "g_exp_ub_is_absolute", False))

    # Build max-expansion capacity path
    kcap_init = data.Kcap_2025 if data.Kcap_2025 else data.Qcap
    kcap_current = {r: float(kcap_init.get(r, data.Qcap.get(r, 0.0))) for r in regions}

    rows = []
    for tp in times:
        for r in regions:
            row = {
                "region": r,
                "t": tp,
                "Q_offer": kcap_current[r],
                "a_bid": float(
                    data.a_dem_t[(r, tp)] if data.a_dem_t else data.a_dem.get(r, 0.0)
                ),
            }
            # p_offer columns: one per importer
            for imp in regions:
                row[f"p_offer_{imp}"] = 0.5 * float(data.p_offer_ub[(r, imp)])

            # dK_net: expansion rate (only for non-terminal periods)
            is_last = tp == times[-1]
            g = float(g_exp_dict.get(r, 0.0))
            k = kcap_current[r]
            rate = g if g_exp_is_abs else g * k
            row["dK_net"] = 0.0 if is_last else rate

            rows.append(row)

        # Roll forward capacity for next period
        if tp != times[-1]:
            for r in regions:
                g = float(g_exp_dict.get(r, 0.0))
                k = kcap_current[r]
                ytn = float(ytn_dict.get(tp, 5.0))
                rate = g if g_exp_is_abs else g * k
                kcap_current[r] = k + ytn * rate

    df = pd.DataFrame(rows)

    # Column ordering: region, t, Q_offer, dK_net, a_bid, p_offer_*
    p_cols = [f"p_offer_{imp}" for imp in regions]
    df = df[["region", "t", "Q_offer", "dK_net", "a_bid"] + p_cols]

    # Write into the existing workbook (preserve all other sheets)
    from openpyxl import load_workbook

    wb = load_workbook(EXCEL_PATH)
    if "initial_state" in wb.sheetnames:
        del wb["initial_state"]
    wb.save(EXCEL_PATH)

    with pd.ExcelWriter(EXCEL_PATH, engine="openpyxl", mode="a") as writer:
        df.to_excel(writer, sheet_name="initial_state", index=False)

    print(f"[OK] Wrote 'initial_state' sheet ({len(df)} rows) to {EXCEL_PATH}")
    print(f"\nPreview:")
    print(df.to_string(index=False))


if __name__ == "__main__":
    main()
